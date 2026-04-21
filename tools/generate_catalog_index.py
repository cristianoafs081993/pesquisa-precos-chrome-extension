import argparse
import glob
import json
import os
import re
import unicodedata

import openpyxl


HEADER_EXTRACTION_ROW = 0
HEADER_ROW = 1
DEFAULT_GLOB = r"C:\Users\crist\Downloads\Cat*Materiais*21112025*.xlsx"

STOPWORDS = {
    "a", "ao", "aos", "as", "com", "da", "das", "de", "do", "dos", "e", "em",
    "na", "nas", "no", "nos", "o", "os", "ou", "para", "por", "um", "uma"
}


def normalize_text(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.strip().lower()


def compact_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def tokenize(text):
    seen = set()
    output = []
    for token in re.split(r"[^a-z0-9]+", normalize_text(text)):
        if len(token) < 3 or token in STOPWORDS or token.isdigit():
            continue
        token = token[:-1] if len(token) > 4 and token.endswith("s") else token
        if token and token not in seen:
            seen.add(token)
            output.append(token)
    return output


def build_search_text(description, pdm):
    text = compact_text(description)
    prefix = compact_text(pdm)
    if prefix and text.upper().startswith(prefix.upper()):
        text = text[len(prefix):].lstrip(" ,;-:")

    text = re.sub(r"\b[\w\u00C0-\u017F /-]{2,40}\s*:\s*", "", text)
    text = re.sub(r"[;,]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    tokens = []
    for source in (pdm, text):
        for token in tokenize(source):
            if token not in tokens:
                tokens.append(token)

    return " ".join(tokens[:18])


def parse_args():
    parser = argparse.ArgumentParser(description="Generate a compact CATMAT catalog index.")
    parser.add_argument("--input", dest="input_path", help="Path to the CATMAT XLSX file.")
    parser.add_argument(
        "--output",
        default=os.path.join("catalog", "catalog-index.json"),
        help="Output path for the generated index.",
    )
    return parser.parse_args()


def resolve_input_path(path_override):
    if path_override:
        return path_override

    matches = glob.glob(DEFAULT_GLOB)
    if not matches:
        raise FileNotFoundError("CATMAT XLSX file not found. Pass --input explicitly.")

    return matches[0]


def main():
    args = parse_args()
    input_path = resolve_input_path(args.input_path)

    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)

    next(rows)  # extraction metadata row
    raw_headers = [compact_text(cell) for cell in next(rows)]
    headers = {normalize_text(cell): index for index, cell in enumerate(raw_headers)}

    records = {}
    row_count = 0

    for row in rows:
        if not any(cell is not None and compact_text(cell) for cell in row):
            continue

        codigo_item = compact_text(row[headers["codigo do item"]])
        if not codigo_item:
            continue

        descricao_item = compact_text(row[headers["descricao do item"]])
        nome_pdm = compact_text(row[headers["nome do pdm"]])
        nome_classe = compact_text(row[headers["nome da classe"]])
        nome_grupo = compact_text(row[headers["nome do grupo"]])
        codigo_ncm = compact_text(row[headers["codigo ncm"]])

        records[codigo_item] = {
            "codigoItem": codigo_item,
            "descricaoItem": descricao_item,
            "nomePdm": nome_pdm,
            "nomeClasse": nome_classe,
            "nomeGrupo": nome_grupo,
            "codigoNcm": "" if codigo_ncm == "-" else codigo_ncm,
            "searchText": build_search_text(descricao_item, nome_pdm),
        }
        row_count += 1

    payload = {
        "version": 1,
        "sourceFile": os.path.basename(input_path),
        "rowCount": row_count,
        "items": dict(sorted(records.items(), key=lambda item: item[0])),
    }

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))

    print(f"Generated {args.output} with {row_count} records from {input_path}")


if __name__ == "__main__":
    main()
